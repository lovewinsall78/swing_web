# app.py
# Streamlit Swing Scanner (KR/US) + 근거(표+차트) + 보유/평단/평가 + 매도추천 + 주문표 엑셀(통화 포맷)
# ------------------------------------------------------------

import re
import io
import math
import numpy as np
import pandas as pd
import yfinance as yf
from pykrx import stock as krx
import streamlit as st
from datetime import datetime, timedelta, date
from typing import Dict, Tuple, Optional
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# =========================
# 기본값
# =========================
DEFAULTS = dict(
    HOLD_DAYS=20,
    MA_FAST=20,
    MA_SLOW=60,
    ATR_PERIOD=14,
    VOL_LOOKBACK=20,
    VOL_SPIKE=1.5,
    ATR_PCT_MIN=0.008,
    ATR_PCT_MAX=0.060,
    ACCOUNT_SIZE=10_000_000,
    RISK_PER_TRADE=0.01,
    STOP_ATR_MULT=1.8,
    TOP_N=10,
    # UI
    FONT_BASE_PX=16,
    FONT_TITLE_PX=26,
)

# 추천(기본 유니버스) - 필요하면 여기만 늘리면 됨
KR_UNIVERSE = [
    "005930", "000660", "035420", "035720", "051910",
    "068270", "005380", "000270", "012330", "028260",
    "105560", "066570", "003670", "034020", "096770",
    "017670", "010130", "009150", "207940", "006400",
]
US_UNIVERSE = [
    "SPY", "QQQ", "IWM", "DIA", "XLK",
    "SMH", "SOXX", "XLF", "XLE", "XLV",
    "AAPL", "MSFT", "NVDA", "AMZN", "GOOGL",
    "META", "TSLA", "BRK-B", "JPM", "UNH",
]

US_NAME_MAP = {
    "SPY": "SPDR S&P 500 ETF",
    "QQQ": "Invesco QQQ (Nasdaq 100)",
    "IWM": "iShares Russell 2000 ETF",
    "DIA": "SPDR Dow Jones ETF",
    "XLK": "Technology Select Sector ETF",
    "SMH": "VanEck Semiconductor ETF",
    "SOXX": "iShares Semiconductor ETF",
    "XLF": "Financial Select Sector ETF",
    "XLE": "Energy Select Sector ETF",
    "XLV": "Health Care Select Sector ETF",
    "BRK-B": "Berkshire Hathaway Class B",
}

# =========================
# 유틸
# =========================
def is_kr_code(x: str) -> bool:
    return bool(re.fullmatch(r"\d{6}", x.strip()))

def normalize_tickers(raw: str):
    items = re.split(r"[,\n\s]+", raw.strip())
    items = [x.strip().upper() for x in items if x.strip()]
    return items

def parse_int_with_commas(s: str, default: int) -> int:
    if s is None:
        return default
    s2 = re.sub(r"[^\d]", "", str(s))
    if s2 == "":
        return default
    try:
        return int(s2)
    except:
        return default

def format_int_with_commas(n: int) -> str:
    try:
        return f"{int(n):,}"
    except:
        return str(n)

def safe_float(v) -> Optional[float]:
    try:
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return None
        if pd.isna(v):
            return None
        return float(v)
    except:
        return None

# =========================
# 지표
# =========================
def compute_atr(df, period=14):
    high, low, close = df["High"], df["Low"], df["Close"]
    prev_close = close.shift(1)
    tr = pd.concat([
        (high - low),
        (high - prev_close).abs(),
        (low - prev_close).abs()
    ], axis=1).max(axis=1)
    return tr.rolling(period).mean()

def add_indicators(df, MA_FAST, MA_SLOW, VOL_LOOKBACK, ATR_PERIOD):
    df = df.copy()
    df["MA_FAST"] = df["Close"].rolling(MA_FAST).mean()
    df["MA_SLOW"] = df["Close"].rolling(MA_SLOW).mean()
    df["VOL_AVG"] = df["Volume"].rolling(VOL_LOOKBACK).mean()
    df["VOL_RATIO"] = np.where(df["VOL_AVG"] > 0, df["Volume"] / df["VOL_AVG"], np.nan)
    df["ATR"] = compute_atr(df, ATR_PERIOD)
    df["ATR_PCT"] = df["ATR"] / df["Close"]
    df["RET_20"] = df["Close"].pct_change(20)
    return df

def rule_signal(last, VOL_SPIKE, ATR_PCT_MIN, ATR_PCT_MAX):
    needed = ["MA_FAST", "MA_SLOW", "VOL_RATIO", "ATR_PCT", "Close"]
    if any(pd.isna(last.get(k, np.nan)) for k in needed):
        return 0
    trend_ok = (last["MA_FAST"] > last["MA_SLOW"]) and (last["Close"] > last["MA_FAST"])
    vol_ok   = (last["VOL_RATIO"] >= VOL_SPIKE)
    atr_ok   = (ATR_PCT_MIN <= last["ATR_PCT"] <= ATR_PCT_MAX)
    return 1 if (trend_ok and vol_ok and atr_ok) else 0

def score_row(last):
    score = 0
    # 추세
    if last.get("MA_FAST", np.nan) > last.get("MA_SLOW", np.nan):
        score += 40
    # 거래량 비율
    vr = last.get("VOL_RATIO", np.nan)
    if pd.notna(vr):
        score += int(min(25, max(0, (vr - 1.0) * 20)))
    # 20일 수익률
    r20 = last.get("RET_20", np.nan)
    if pd.notna(r20):
        score += int(min(20, max(0, r20 * 200)))
    # 변동성(적당하면 +, 너무 크거나 너무 작으면 -)
    ap = last.get("ATR_PCT", np.nan)
    if pd.notna(ap):
        if ap > 0.045:
            score -= 10
        elif ap < 0.010:
            score -= 5
        else:
            score += 10
    return int(score)

def position_size(entry, atr, ACCOUNT_SIZE, RISK_PER_TRADE, STOP_ATR_MULT):
    if pd.isna(atr) or atr is None or atr <= 0:
        return 0, np.nan
    stop_dist = STOP_ATR_MULT * atr
    risk_amount = ACCOUNT_SIZE * RISK_PER_TRADE
    qty = int(risk_amount / stop_dist)
    stop_price = entry - stop_dist
    return max(0, qty), float(stop_price)

def build_reason_table(last, params) -> pd.DataFrame:
    close = safe_float(last.get("Close", np.nan))
    ma_fast = safe_float(last.get("MA_FAST", np.nan))
    ma_slow = safe_float(last.get("MA_SLOW", np.nan))
    vol_ratio = safe_float(last.get("VOL_RATIO", np.nan))
    atr_pct = safe_float(last.get("ATR_PCT", np.nan))

    rows = []

    c1 = (ma_fast is not None) and (ma_slow is not None) and (ma_fast > ma_slow)
    c2 = (close is not None) and (ma_fast is not None) and (close > ma_fast)
    rows.append({
        "조건": "추세: MA_FAST > MA_SLOW",
        "현재값": f"{ma_fast:.2f} > {ma_slow:.2f}" if (ma_fast is not None and ma_slow is not None) else "데이터 부족",
        "기준": "단기선이 장기선 위",
        "통과": "✅" if c1 else "❌"
    })
    rows.append({
        "조건": "추세: Close > MA_FAST",
        "현재값": f"{close:.2f} > {ma_fast:.2f}" if (close is not None and ma_fast is not None) else "데이터 부족",
        "기준": "종가가 단기선 위",
        "통과": "✅" if c2 else "❌"
    })

    vol_spike = float(params["VOL_SPIKE"])
    c3 = (vol_ratio is not None) and (vol_ratio >= vol_spike)
    rows.append({
        "조건": "거래량: VOL_RATIO ≥ VOL_SPIKE",
        "현재값": f"{vol_ratio:.2f}" if (vol_ratio is not None) else "데이터 부족",
        "기준": f"≥ {vol_spike:.2f}",
        "통과": "✅" if c3 else "❌"
    })

    atr_min = float(params["ATR_PCT_MIN"])
    atr_max = float(params["ATR_PCT_MAX"])
    c4 = (atr_pct is not None) and (atr_min <= atr_pct <= atr_max)
    rows.append({
        "조건": "변동성: ATR% 범위",
        "현재값": f"{atr_pct*100:.2f}%" if (atr_pct is not None) else "데이터 부족",
        "기준": f"{atr_min*100:.2f}% ~ {atr_max*100:.2f}%",
        "통과": "✅" if c4 else "❌"
    })

    return pd.DataFrame(rows)

# =========================
# 데이터 로드 (KR/US)
# =========================
@st.cache_data(ttl=60*30, show_spinner=False)
def get_kr_name(code: str) -> str:
    try:
        return krx.get_market_ticker_name(code)
    except:
        return ""

@st.cache_data(ttl=60*60, show_spinner=False)
def get_us_name(ticker: str) -> str:
    if ticker in US_NAME_MAP:
        return US_NAME_MAP[ticker]
    try:
        info = yf.Ticker(ticker).info
        return info.get("shortName") or info.get("longName") or ""
    except:
        return ""

@st.cache_data(ttl=60*20, show_spinner=False)
def load_us(ticker: str) -> pd.DataFrame:
    df = yf.download(
        ticker,
        period="2y",
        interval="1d",
        auto_adjust=False,
        progress=False,
        group_by="column",
        threads=False,
    )

    # MultiIndex 안전 처리
    if isinstance(df.columns, pd.MultiIndex):
        lv0 = df.columns.get_level_values(0)
        lv1 = df.columns.get_level_values(1)

        if ticker in lv0:
            df = df[ticker]
        elif ticker in lv1:
            df = df.xs(ticker, axis=1, level=1)
        else:
            # 최후의 수단: 첫 번째 티커만
            uniq0 = list(pd.unique(lv0))
            uniq1 = list(pd.unique(lv1))
            if len(uniq0) > 0 and len(uniq1) > 0:
                # level1이 티커일 확률이 높으면 level1 첫번째
                if any(x.upper() == ticker for x in uniq1):
                    df = df.xs(uniq1[0], axis=1, level=1)
                else:
                    df = df[uniq0[0]]
            else:
                raise ValueError("Unexpected MultiIndex columns")

    # 컬럼명 title() 적용은 '문자열 컬럼명'에만
    df.columns = [str(c).title() for c in df.columns]

    # Close 없고 Adj Close만 있으면 대체
    if "Close" not in df.columns and "Adj Close" in df.columns:
        df["Close"] = df["Adj Close"]

    keep = ["Open", "High", "Low", "Close", "Volume"]
    missing = [c for c in keep if c not in df.columns]
    if missing:
        raise ValueError(f"US data missing columns: {missing} / columns={list(df.columns)}")

    df = df[keep].dropna()
    return df

@st.cache_data(ttl=60*30, show_spinner=False)
def load_kr(code: str) -> pd.DataFrame:
    end = datetime.now().strftime("%Y-%m-%d")
    start = (datetime.now() - timedelta(days=365*2)).strftime("%Y-%m-%d")
    df = krx.get_market_ohlcv_by_date(start, end, code)
    df = df.rename(columns={"시가":"Open","고가":"High","저가":"Low","종가":"Close","거래량":"Volume"})
    df = df[["Open","High","Low","Close","Volume"]].dropna()
    return df

# =========================
# 분석
# =========================
def analyze_one(ticker: str, params: dict) -> Tuple[dict, Optional[pd.DataFrame]]:
    """
    return: (row_dict, df_with_indicators or None)
    """
    try:
        if is_kr_code(ticker):
            market = "KR"
            name = get_kr_name(ticker)
            df = load_kr(ticker)
        else:
            market = "US"
            name = get_us_name(ticker)
            df = load_us(ticker)

        df = add_indicators(
            df,
            MA_FAST=params["MA_FAST"],
            MA_SLOW=params["MA_SLOW"],
            VOL_LOOKBACK=params["VOL_LOOKBACK"],
            ATR_PERIOD=params["ATR_PERIOD"],
        )
        last = df.iloc[-1]

        cand = rule_signal(
            last,
            VOL_SPIKE=params["VOL_SPIKE"],
            ATR_PCT_MIN=params["ATR_PCT_MIN"],
            ATR_PCT_MAX=params["ATR_PCT_MAX"],
        )
        sc = score_row(last) if cand == 1 else 0

        entry = float(last["Close"])
        qty, stop = position_size(
            entry, float(last["ATR"]),
            ACCOUNT_SIZE=params["ACCOUNT_SIZE"],
            RISK_PER_TRADE=params["RISK_PER_TRADE"],
            STOP_ATR_MULT=params["STOP_ATR_MULT"],
        )
        target = entry + 2 * (entry - stop) if (not np.isnan(stop)) else np.nan

        row = {
            "market": market,
            "ticker": ticker,
            "name": name,
            "O/X": "O" if cand == 1 else "X",
            "candidate": int(cand),
            "score": int(sc),
            "close": round(entry, 2),
            "stop": (None if np.isnan(stop) else round(stop, 2)),
            "target(2R)": (None if np.isnan(target) else round(target, 2)),
            "qty": int(qty),
            "vol_ratio": (None if pd.isna(last["VOL_RATIO"]) else round(float(last["VOL_RATIO"]), 2)),
            "atr_pct(%)": (None if pd.isna(last["ATR_PCT"]) else round(float(last["ATR_PCT"])*100, 2)),
            "ret_20(%)": (None if pd.isna(last["RET_20"]) else round(float(last["RET_20"])*100, 2)),
            "date": str(df.index[-1].date()),
            "error": ""
        }
        return row, df
    except Exception as e:
        row = {
            "market": "?",
            "ticker": ticker,
            "name": "",
            "O/X": "X",
            "candidate": 0,
            "score": 0,
            "close": None,
            "stop": None,
            "target(2R)": None,
            "qty": 0,
            "vol_ratio": None,
            "atr_pct(%)": None,
            "ret_20(%)": None,
            "date": "",
            "error": str(e)
        }
        return row, None

# =========================
# 보유/매도 추천
# =========================
def to_date(x) -> Optional[date]:
    if x is None or x == "":
        return None
    if isinstance(x, date):
        return x
    if isinstance(x, datetime):
        return x.date()
    try:
        return pd.to_datetime(x).date()
    except:
        return None

def sell_recommend(pos_row: dict, last: pd.Series, params: dict) -> Tuple[str, str]:
    """
    return (action, reason)
    action: HOLD / SELL / WATCH
    """
    close = safe_float(last.get("Close", np.nan))
    ma_fast = safe_float(last.get("MA_FAST", np.nan))
    ma_slow = safe_float(last.get("MA_SLOW", np.nan))
    stop = safe_float(pos_row.get("stop_calc", np.nan))  # 계산된 손절(보유기준)
    buy_date = to_date(pos_row.get("buy_date"))
    hold_days = None
    if buy_date is not None:
        hold_days = (date.today() - buy_date).days

    # 1) 손절
    if close is not None and stop is not None and close < stop:
        return "SELL", "손절가 하회"

    # 2) 추세 훼손
    if ma_fast is not None and ma_slow is not None and ma_fast < ma_slow:
        return "SELL", "단기선 < 장기선(추세 약화)"

    # 3) 최대 보유일 초과
    if hold_days is not None and hold_days >= int(params["HOLD_DAYS"]):
        return "SELL", f"최대 보유일({int(params['HOLD_DAYS'])}일) 초과"

    # 4) 그 외
    if close is not None and stop is not None:
        # stop에 근접하면 WATCH
        if (close - stop) / max(close, 1e-9) < 0.02:
            return "WATCH", "손절가 근접(주의)"
    return "HOLD", "유지"

# =========================
# 엑셀 생성 (통화 포맷 포함)
# =========================
def autosize_columns(ws, max_width=44):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_width, max(10, max_len + 2))

def apply_currency_formats(ws, df: pd.DataFrame, market_col: str, cols_money: list):
    # 헤더 찾기
    headers = [c.value for c in ws[1]]
    idx = {h: i+1 for i, h in enumerate(headers)}

    for r in range(2, ws.max_row + 1):
        mkt = ws.cell(r, idx[market_col]).value if market_col in idx else ""
        for c_name in cols_money:
            if c_name not in idx:
                continue
            cell = ws.cell(r, idx[c_name])
            if cell.value is None or cell.value == "":
                continue
            # 시장별 포맷
            if mkt == "KR":
                # 원화: 정수 느낌 (필요시 소수점 제거)
                cell.number_format = '#,##0'
            elif mkt == "US":
                cell.number_format = '$#,##0.00'
            else:
                cell.number_format = '#,##0.00'

def build_excel(df_all: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 1) 전체
        df_all.to_excel(writer, sheet_name="Signals_All", index=False)

        # 2) 후보만
        df_cand = df_all[df_all["candidate"] == 1].copy()
        if df_cand.empty:
            df_cand = pd.DataFrame([{"note": "nottoday"}])
        df_cand.to_excel(writer, sheet_name="Candidates", index=False)

        # 3) 주문표
        def order_sheet(df, market):
            d = df[(df["market"] == market) & (df["candidate"] == 1)].copy()
            if d.empty:
                return pd.DataFrame([{
                    "market": market, "ticker": "nottoday", "name": "",
                    "action": "NONE", "qty": 0, "stop": "", "target(2R)": "",
                    "note": "No candidates"
                }])
            d = d.sort_values("score", ascending=False).head(10)
            d["qty_safe"] = (d["qty"] * 0.5).astype(int)
            return pd.DataFrame({
                "market": d["market"],
                "ticker": d["ticker"],
                "name": d.get("name", ""),
                "action": "BUY",
                "qty": d["qty_safe"],
                "stop": d["stop"],
                "target(2R)": d["target(2R)"],
                "note": ["Manual entry (M-STOCK)"] * len(d)
            })

        order_sheet(df_all, "KR").to_excel(writer, sheet_name="Order_KR", index=False)
        order_sheet(df_all, "US").to_excel(writer, sheet_name="Order_US", index=False)

    # openpyxl로 포맷 적용
    output.seek(0)
    from openpyxl import load_workbook
    wb = load_workbook(output)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.freeze_panes = "A2"
        # 헤더 스타일
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        autosize_columns(ws)

    # 통화 포맷: Signals_All, Candidates, Order_*에 적용
    if "Signals_All" in wb.sheetnames:
        ws = wb["Signals_All"]
        apply_currency_formats(ws, df_all, "market", ["close", "stop", "target(2R)"])
    if "Candidates" in wb.sheetnames:
        ws = wb["Candidates"]
        # Candidates는 컬럼명이 다를 수 있음
        headers = [c.value for c in ws[1]]
        if "market" in headers:
            apply_currency_formats(ws, df_all, "market", ["close", "stop", "target(2R)"])
    for s in ["Order_KR", "Order_US"]:
        if s in wb.sheetnames:
            ws = wb[s]
            # Order sheet는 시장이 한 종류지만 통일 포맷
            headers = [c.value for c in ws[1]]
            idx = {h: i+1 for i, h in enumerate(headers)}
            for r in range(2, ws.max_row + 1):
                for c_name in ["stop", "target(2R)"]:
                    if c_name in idx:
                        cell = ws.cell(r, idx[c_name])
                        if cell.value is None or cell.value == "":
                            continue
                        if s.endswith("KR"):
                            cell.number_format = '#,##0'
                        else:
                            cell.number_format = '$#,##0.00'

    out2 = io.BytesIO()
    wb.save(out2)
    return out2.getvalue()

# =========================
# UI: 폰트(방법1 - CSS)
# =========================
def inject_css(font_base_px: int, font_title_px: int):
    st.markdown(
        f"""
        <style>
        html, body, [class*="css"]  {{
            font-size: {font_base_px}px !important;
        }}
        h1 {{
            font-size: {font_title_px}px !important;
        }}
        /* 데이터프레임 가로 가시성 */
        .stDataFrame, .stDataEditor {{
            width: 100% !important;
        }}
        </style>
        """,
        unsafe_allow_html=True
    )

# =========================
# 세션 초기화(리셋 방지 핵심)
# =========================
def init_state():
    if "ticker_input" not in st.session_state:
        st.session_state["ticker_input"] = "005930 000660\nSPY QQQ"
    if "positions_df" not in st.session_state:
        st.session_state["positions_df"] = pd.DataFrame([
            {"market": "KR", "ticker": "005930", "name": get_kr_name("005930"), "qty": 0, "avg_price": 0, "buy_date": ""},
            {"market": "US", "ticker": "SPY",    "name": get_us_name("SPY"),    "qty": 0, "avg_price": 0, "buy_date": ""},
        ])
    if "account_size_text" not in st.session_state:
        st.session_state["account_size_text"] = format_int_with_commas(DEFAULTS["ACCOUNT_SIZE"])
    if "font_base_px" not in st.session_state:
        st.session_state["font_base_px"] = DEFAULTS["FONT_BASE_PX"]
    if "font_title_px" not in st.session_state:
        st.session_state["font_title_px"] = DEFAULTS["FONT_TITLE_PX"]

# =========================
# 페이지
# =========================
st.set_page_config(page_title="Swing Scanner", layout="wide")
init_state()

# 폰트 적용(방법1)
inject_css(st.session_state["font_base_px"], st.session_state["font_title_px"])

st.title("웹 티커 입력 → 스윙 판단(O/X) + 근거(표+차트) + 보유/매도 + 주문표 엑셀")

# =========================
# 사이드바(카테고리 구분 + 설명)
# =========================
with st.sidebar:
    st.header("스윙 전략 설정")

    # 폰트
    st.markdown("### 화면 폰트")
    st.session_state["font_base_px"] = st.slider("기본 글자 크기(px)", 12, 22, int(st.session_state["font_base_px"]), step=1)
    st.session_state["font_title_px"] = st.slider("제목 글자 크기(px)", 18, 40, int(st.session_state["font_title_px"]), step=1)
    st.caption("※ 폰트는 CSS로 적용(방법1). 조절 후 화면은 즉시 반영됩니다.")

    st.markdown("---")
    st.markdown("### 1) 추세(이동평균)")
    params = {}
    params["MA_FAST"] = st.number_input("MA_FAST (단기 이동평균)", 5, 200, DEFAULTS["MA_FAST"], key="MA_FAST")
    st.write("단기 흐름을 보는 평균. 작을수록 신호가 빠르지만 잦은 실패 가능.")
    params["MA_SLOW"] = st.number_input("MA_SLOW (장기 이동평균)", 10, 300, DEFAULTS["MA_SLOW"], key="MA_SLOW")
    st.write("중·장기 추세 기준. MA_FAST가 MA_SLOW 위면 상승추세로 봅니다.")

    st.markdown("---")
    st.markdown("### 2) 거래량/변동성")
    params["VOL_LOOKBACK"] = st.number_input("VOL_LOOKBACK (거래량 평균 기간)", 5, 200, DEFAULTS["VOL_LOOKBACK"], key="VOL_LOOKBACK")
    st.write("평균 거래량 계산 기간.")
    params["VOL_SPIKE"] = st.number_input("VOL_SPIKE (거래량 급증 기준)", 1.0, 5.0, float(DEFAULTS["VOL_SPIKE"]), step=0.05, key="VOL_SPIKE")
    st.write("현재 거래량이 평균 대비 몇 배 이상인지. 예: 1.5 = 150% 이상.")
    params["ATR_PERIOD"] = st.number_input("ATR_PERIOD (ATR 계산 기간)", 5, 100, DEFAULTS["ATR_PERIOD"], key="ATR_PERIOD")
    st.write("ATR은 평균 변동폭. 값이 클수록 변동이 큼.")
    params["ATR_PCT_MIN"] = st.number_input("ATR_PCT_MIN (최소 변동성)", 0.0, 0.2, float(DEFAULTS["ATR_PCT_MIN"]), step=0.001, format="%.3f", key="ATR_PCT_MIN")
    params["ATR_PCT_MAX"] = st.number_input("ATR_PCT_MAX (최대 변동성)", 0.0, 0.5, float(DEFAULTS["ATR_PCT_MAX"]), step=0.001, format="%.3f", key="ATR_PCT_MAX")
    st.write("ATR%가 너무 작으면(안 움직임) 제외, 너무 크면(고위험) 제외.")

    st.markdown("---")
    st.markdown("### 3) 자금/리스크")
    # 계좌 가정값(천단위 콤마 가능) - text_input 사용
    st.session_state["account_size_text"] = st.text_input(
        "ACCOUNT_SIZE (총 투자금, 콤마 가능)",
        value=st.session_state["account_size_text"],
        key="ACCOUNT_SIZE_TEXT",
        help="예: 10,000,000"
    )
    params["ACCOUNT_SIZE"] = parse_int_with_commas(st.session_state["account_size_text"], DEFAULTS["ACCOUNT_SIZE"])
    st.caption(f"인식된 값: {format_int_with_commas(params['ACCOUNT_SIZE'])}")

    params["RISK_PER_TRADE"] = st.number_input("RISK_PER_TRADE (1회 최대 손실 비율)", 0.001, 0.05, float(DEFAULTS["RISK_PER_TRADE"]), step=0.001, format="%.3f", key="RISK_PER_TRADE")
    st.write("예: 0.01 = 1% 손실 한도.")
    params["STOP_ATR_MULT"] = st.number_input("STOP_ATR_MULT (손절 ATR 배수)", 0.5, 5.0, float(DEFAULTS["STOP_ATR_MULT"]), step=0.1, key="STOP_ATR_MULT")
    st.write("손절 여유폭. 보통 1.5~2.0 자주 사용.")
    params["HOLD_DAYS"] = st.number_input("HOLD_DAYS (최대 보유일)", 1, 200, DEFAULTS["HOLD_DAYS"], key="HOLD_DAYS")
    st.write("보유일이 이 값을 넘으면 매도(또는 청산) 후보로 표시.")

# 폰트 슬라이더 반영
inject_css(st.session_state["font_base_px"], st.session_state["font_title_px"])

# =========================
# 기능 확인 리스트 (요청하신 “포함 기능 확인”)
# =========================
with st.expander("✅ 포함된 기능 확인 리스트(누락 방지 체크)", expanded=True):
    st.write(
        "- KR/US 티커 입력(텍스트박스, 자동 채움 가능)\n"
        "- 스윙 후보(O/X) 판정 + 점수(score)\n"
        "- 근거 표(조건별 통과/실패) + 차트(종가/이평)\n"
        "- 추천 스캔: KR/US/균형(5+5) 옵션, 추천 후 티커 입력창 자동 반영\n"
        "- 회사명(name) 표시(KR=pykrx, US=가능한 범위에서 yfinance/내장 맵)\n"
        "- 보유 포지션 입력(수량/평단/매수일) + 평가/손익 자동 계산\n"
        "- 매도 추천(HOLD/SELL/WATCH) + 이유\n"
        "- 엑셀 다운로드(Signals/Candidates/Order_KR/Order_US) + 통화 포맷(KRW 콤마, US 달러)\n"
        "- 입력 리셋 방지(세션 상태/에디터 키 분리)\n"
    )

st.write("**입력 방법**: KR은 6자리(예: `005930`), US는 티커(예: `SPY`). 콤마/줄바꿈/공백 모두 가능.")

# =========================
# 추천 스캔(가시성: 티커 입력창 근처)
# =========================
st.subheader("추천 스캔(전략 기준으로 상위 N 자동 추천)")
c1, c2, c3, c4 = st.columns([1.1, 1.1, 1.1, 1.2])
with c1:
    rec_mode = st.selectbox("추천 옵션", ["균형추천(5+5)", "KR만", "US만"], index=0)
with c2:
    rec_topn = st.number_input("추천 개수(N)", 5, 30, 10, step=1)
with c3:
    run_rec = st.button("추천 실행", use_container_width=True)
with c4:
    st.caption("※ 추천은 미리 정한 유니버스에서 점수(score) 상위로 뽑습니다.")

def scan_recommend(universe: list, params: dict) -> pd.DataFrame:
    rows = []
    for t in universe:
        row, _ = analyze_one(t, params)
        if row.get("error"):
            continue
        rows.append(row)
    if not rows:
        return pd.DataFrame()
    dfu = pd.DataFrame(rows)
    # 후보 우선, 그 다음 score
    dfu = dfu.sort_values(["candidate", "score"], ascending=[False, False]).reset_index(drop=True)
    return dfu

rec_table = None
if run_rec:
    if rec_mode == "KR만":
        dfu = scan_recommend(KR_UNIVERSE, params)
        rec_table = dfu.head(int(rec_topn))
    elif rec_mode == "US만":
        dfu = scan_recommend(US_UNIVERSE, params)
        rec_table = dfu.head(int(rec_topn))
    else:
        dfk = scan_recommend(KR_UNIVERSE, params)
        dfu = scan_recommend(US_UNIVERSE, params)
        k_take = max(1, int(rec_topn) // 2)
        u_take = int(rec_topn) - k_take
        rec_table = pd.concat([dfk.head(k_take), dfu.head(u_take)], ignore_index=True)

    if rec_table is not None and not rec_table.empty:
        show_cols = ["market", "ticker", "name", "O/X", "score", "close", "vol_ratio", "atr_pct(%)", "ret_20(%)", "date"]
        st.dataframe(rec_table[show_cols], use_container_width=True)

        # ✅ 추천 후 티커 입력창 자동 반영(핵심)
        tickers_fill = "\n".join(rec_table["ticker"].astype(str).tolist())
        st.session_state["ticker_input"] = tickers_fill
        st.success("추천 티커를 아래 '티커 입력' 칸에 자동으로 넣었습니다.")
    else:
        st.error("추천 결과가 비었습니다(데이터/유니버스/네트워크 확인).")

st.markdown("---")

# =========================
# 티커 입력(세션키로 리셋 방지)
# =========================
raw = st.text_area("티커 입력", key="ticker_input", height=120)

run = st.button("분석 실행", type="primary")

# =========================
# 보유 포지션(리셋/2번 입력 문제 해결)
# - data_editor key와 session_state 저장 key를 분리!
# =========================
st.subheader("보유 포지션 입력(평단/수량/매수일) → 평가/손익/매도추천 자동")
st.caption("평단/수량/매수일을 입력하면, 종가를 불러와 평가금/손익/매도추천이 계산됩니다.")

positions_src = st.session_state["positions_df"].copy()

edited_positions = st.data_editor(
    positions_src,
    key="positions_editor",  # 위젯 키(절대 session_state에 직접 값 대입 금지)
    use_container_width=True,
    num_rows="dynamic",
    column_config={
        "market": st.column_config.SelectboxColumn("market", options=["KR", "US"], required=True),
        "ticker": st.column_config.TextColumn("ticker", width="small"),
        "name": st.column_config.TextColumn("회사명", width="medium"),
        "qty": st.column_config.NumberColumn("보유수량", min_value=0, step=1),
        "avg_price": st.column_config.NumberColumn("평단(진입가)", min_value=0.0, step=0.01),
        "buy_date": st.column_config.TextColumn("매수일(YYYY-MM-DD)", width="small"),
    },
)

# ✅ 편집 결과는 다른 키에 저장(리셋/2번 입력 문제 해결 핵심)
st.session_state["positions_df"] = edited_positions.copy()

# =========================
# 분석 실행
# =========================
if run:
    tickers = normalize_tickers(raw)
    if not tickers:
        st.warning("티커를 1개 이상 입력하세요.")
        st.stop()

    # 전체 분석
    rows = []
    df_map = {}  # 근거/차트용
    for t in tickers:
        row, df_ind = analyze_one(t, params)
        rows.append(row)
        if df_ind is not None and row.get("error") == "":
            df_map[t] = df_ind

    df = pd.DataFrame(rows)
    df = df.sort_values(["candidate", "score"], ascending=[False, False]).reset_index(drop=True)

    st.subheader("분석 결과(가로 가시성 최대)")
    st.dataframe(df, use_container_width=True)

    n_cand = int((df["candidate"] == 1).sum())
    if n_cand > 0:
        st.success(f"후보(O) {n_cand}개")
    else:
        st.error("NOT TODAY: 조건을 만족하는 후보가 없습니다. (O=0)")

    # =========================
    # 근거(표+차트)
    # =========================
    st.subheader("근거(조건표 + 차트)")
    st.caption("각 티커별로 왜 O/X인지 조건별 통과 여부를 표로 보여주고, 종가/이평 차트를 함께 표시합니다.")

    for _, r in df.iterrows():
        t = str(r["ticker"])
        mkt = r.get("market", "?")
        name = r.get("name", "")
        ox = r.get("O/X", "X")
        score = r.get("score", 0)
        err = r.get("error", "")

        title = f"[{mkt}] {t}  {name}  |  O/X={ox}  score={score}"
        with st.expander(title, expanded=False):
            if err:
                st.error(f"데이터 오류: {err}")
                continue
            if t not in df_map:
                st.warning("차트/근거를 만들 데이터가 없습니다.")
                continue

            df_ind = df_map[t]
            last = df_ind.iloc[-1]

            reason_df = build_reason_table(last, params)
            st.dataframe(reason_df, use_container_width=True)

            # 차트(간단·가시성 위주)
            chart_df = df_ind[["Close", "MA_FAST", "MA_SLOW"]].dropna().copy()
            chart_df = chart_df.tail(180)
            st.line_chart(chart_df, use_container_width=True)

            st.caption(
                "차트 해석: Close가 MA_FAST 위 + MA_FAST가 MA_SLOW 위면 추세 조건 통과 가능성이 큽니다.\n"
                "거래량/ATR% 조건은 위 표에서 함께 확인하세요."
            )

    # =========================
    # 보유 포지션 평가/매도 추천 계산
    # =========================
    st.subheader("보유 포지션 평가/매도 추천(자동 계산)")
    pos_df = st.session_state["positions_df"].copy()
    if pos_df.empty:
        st.info("보유 포지션이 없습니다.")
    else:
        # ticker/name 자동 보정
        def fill_name(row):
            m = row.get("market", "")
            t = str(row.get("ticker", "")).strip().upper()
            if not t:
                return ""
            if m == "KR" and is_kr_code(t):
                return get_kr_name(t)
            if m == "US":
                return get_us_name(t)
            return row.get("name", "")

        pos_df["ticker"] = pos_df["ticker"].astype(str).str.strip().str.upper()
        pos_df["name"] = pos_df.apply(fill_name, axis=1)

        # 계산 컬럼
        last_close = []
        stop_calc = []
        mkt_value = []
        pnl = []
        pnl_pct = []
        hold_days_list = []
        action_list = []
        reason_list = []

        # 포지션마다 분석(df_ind) 재사용: df_map에 없으면 새로 로드/지표
        for _, pr in pos_df.iterrows():
            m = pr.get("market", "")
            t = str(pr.get("ticker", "")).strip().upper()
            qty = safe_float(pr.get("qty", 0)) or 0
            avgp = safe_float(pr.get("avg_price", 0)) or 0
            bd = to_date(pr.get("buy_date"))

            if not t:
                last_close.append(np.nan)
                stop_calc.append(np.nan)
                mkt_value.append(np.nan)
                pnl.append(np.nan)
                pnl_pct.append(np.nan)
                hold_days_list.append("")
                action_list.append("")
                reason_list.append("")
                continue

            # df_ind 확보
            df_ind = df_map.get(t)
            if df_ind is None:
                # 포지션 티커가 입력 티커에 없을 수도 있으니 별도 로드
                try:
                    base = load_kr(t) if (m == "KR" and is_kr_code(t)) else load_us(t)
                    df_ind = add_indicators(
                        base,
                        MA_FAST=params["MA_FAST"],
                        MA_SLOW=params["MA_SLOW"],
                        VOL_LOOKBACK=params["VOL_LOOKBACK"],
                        ATR_PERIOD=params["ATR_PERIOD"],
                    )
                except:
                    df_ind = None

            if df_ind is None or df_ind.empty:
                last_close.append(np.nan)
                stop_calc.append(np.nan)
                mkt_value.append(np.nan)
                pnl.append(np.nan)
                pnl_pct.append(np.nan)
                hold_days_list.append("")
                action_list.append("WATCH")
                reason_list.append("데이터 로드 실패")
                continue

            last = df_ind.iloc[-1]
            c = safe_float(last.get("Close", np.nan))
            atr = safe_float(last.get("ATR", np.nan))

            # 보유 기준 손절(평단 - STOP_ATR_MULT*ATR)
            if avgp > 0 and atr is not None and atr > 0:
                stop_p = avgp - float(params["STOP_ATR_MULT"]) * atr
            else:
                stop_p = np.nan

            # 평가/손익
            if c is not None:
                mv = qty * c
                p = qty * (c - avgp) if avgp > 0 else np.nan
                pp = ((c - avgp) / avgp * 100.0) if avgp > 0 else np.nan
            else:
                mv = np.nan
                p = np.nan
                pp = np.nan

            # 보유일
            hd = (date.today() - bd).days if bd is not None else None

            # 매도 추천
            tmp_row = dict(pr)
            tmp_row["stop_calc"] = stop_p
            act, rsn = sell_recommend(tmp_row, last, params)

            last_close.append(c if c is not None else np.nan)
            stop_calc.append(stop_p)
            mkt_value.append(mv)
            pnl.append(p)
            pnl_pct.append(pp)
            hold_days_list.append(hd if hd is not None else "")
            action_list.append(act)
            reason_list.append(rsn)

        pos_df["last_close"] = last_close
        pos_df["stop_calc"] = stop_calc
        pos_df["mkt_value"] = mkt_value
        pos_df["pnl"] = pnl
        pos_df["pnl_pct(%)"] = pnl_pct
        pos_df["hold_days"] = hold_days_list
        pos_df["action"] = action_list
        pos_df["action_reason"] = reason_list

        # 통화표시용: KR/US 구분해서 보여주기(표시는 문자열로)
        def fmt_money(mkt, v):
            if v is None or (isinstance(v, float) and np.isnan(v)):
                return ""
            if mkt == "KR":
                return f"{float(v):,.0f}원"
            if mkt == "US":
                return f"${float(v):,.2f}"
            return f"{float(v):,.2f}"

        show = pos_df.copy()
        show["last_close_disp"] = show.apply(lambda r: fmt_money(r["market"], r["last_close"]), axis=1)
        show["stop_calc_disp"] = show.apply(lambda r: fmt_money(r["market"], r["stop_calc"]), axis=1)
        show["mkt_value_disp"] = show.apply(lambda r: fmt_money(r["market"], r["mkt_value"]), axis=1)
        show["pnl_disp"] = show.apply(lambda r: fmt_money(r["market"], r["pnl"]), axis=1)

        view_cols = ["market", "ticker", "name", "qty", "avg_price", "buy_date",
                    "last_close_disp", "stop_calc_disp", "mkt_value_disp", "pnl_disp", "pnl_pct(%)",
                    "hold_days", "action", "action_reason"]
        st.dataframe(show[view_cols], use_container_width=True)

        st.caption(
            "매도 추천 기준(요약):\n"
            "- 종가가 손절가(stop_calc) 아래면 SELL\n"
            "- MA_FAST < MA_SLOW면 SELL\n"
            "- 보유일이 HOLD_DAYS 이상이면 SELL\n"
            "- 손절가 근접(2% 이내)이면 WATCH\n"
        )

    # =========================
    # 엑셀 다운로드 (누락 방지)
    # =========================
    st.subheader("엑셀 다운로드")
    xlsx_bytes = build_excel(df)
    st.download_button(
        label="엑셀 1개로 다운로드 (All-in-One)",
        data=xlsx_bytes,
        file_name="Swing_Output_AllInOne.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # =========================
    # 다음 단계 3가지(요청 반영)
    # =========================
    st.subheader("다음 단계(3가지)")
    st.write(
        "1) **유니버스 확장**: KR/US 후보군을 50~200개로 늘려 추천 품질을 높이기\n"
        "2) **매도 룰 고도화**: 익절(트레일링/2R 도달/MA 이탈) 규칙을 옵션으로 추가\n"
        "3) **백테스트 탭 추가**: 최근 2년 기준으로 후보 신호의 성과(승률/손익비)를 요약 테이블로 제공\n"
    )
